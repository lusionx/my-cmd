# -*- coding: utf-8 -*-
#!/usr/bin/python


from openpyxl.reader.excel import load_workbook
from models import *
import os

constr = 'sqlite:///'+os.path.dirname(__file__)+'/info.sqlite3'

def main(path,sheetindex = 0,):
    column = u'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    wb = load_workbook(filename = path)
    sheet = wb.worksheets[sheetindex]
    ss = Context(constr).session
    for i in xrange(2,30):
        a = Carton()
        a.name = sheet.cell('A'+str(i)).value
        a.week = sheet.cell('B'+str(i)).value
        a.start = '2011-'+sheet.cell('C'+str(i)).value

    print sheet.cell('D6').value#.encode('utf-8')
    

if __name__ == '__main__':
    main(u'D:/admin/Desktop/新番动漫.xlsx')
